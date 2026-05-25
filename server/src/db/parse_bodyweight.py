#!/usr/bin/env python3
"""Parser for Bodyweight Zone.xlsx.

Each workout block repeats a header row whose cells name the columns
(TIME | ORDER | FOCUS | LEVEL | EXERCISE | REPS). We read the column positions
from that header, treat each header occurrence as a new workout, and take the
labelled rows (A1/B1...) as the exercises. Unlabelled LVL2/LVL3 continuation
rows are progressions of the slot above and are skipped. There is no SETS
column, so sets default to 1.

  --dump "<sheet>" | --import
"""

import os
import re
import sys
import sqlite3

import openpyxl

DB_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'ageless.db')
DATA_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'import-data')
FILE = 'Bodyweight Zone.xlsx'
PREFIX = 'Bodyweight'

LABEL_RE = re.compile(r'^([A-Z])(\d+)$')
LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
MAXCOL = 12


def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return '' if v is None else re.sub(r'\s+', ' ', str(v)).strip()


def parse_reps(raw):
    out = {'reps': None, 'duration_secs': None, 'time_based': 0, 'per_side': None, 'tracking_type': None}
    if not raw:
        return out
    val = raw.strip()
    if re.search(r'e\s*[./]\s*s|each side|per side|partnered', val, re.IGNORECASE):
        out['per_side'] = 'side'
    clean = re.sub(r'\s*(e\s*[./]\s*s\.?|each side|per side)', '', val, flags=re.IGNORECASE).strip()
    mt = re.match(r'^(\d+):(\d{2})$', clean)
    ms = re.match(r'^(\d+)\s*s$', clean, re.IGNORECASE)
    if mt:
        out.update(duration_secs=int(mt.group(1)) * 60 + int(mt.group(2)), time_based=1, tracking_type='Duration', reps=clean)
        return out
    if ms:
        out.update(duration_secs=int(ms.group(1)), time_based=1, tracking_type='Duration', reps=clean)
        return out
    mr = re.match(r'^(\d+(?:\s*-\s*\d+)?)\s*r', clean, re.IGNORECASE)
    if mr:
        out.update(reps=re.sub(r'\s*', '', mr.group(1)), tracking_type='Repetitions')
        return out
    out['reps'] = clean
    return out


def header_cols(ws, r):
    """If row r is a column-header row, return {order, exercise, reps} column indices."""
    found = {}
    for c in range(1, MAXCOL + 1):
        v = cell(ws, r, c).upper()
        if v == 'ORDER':
            found['order'] = c
        elif v == 'EXERCISE':
            found['exercise'] = c
        elif v == 'REPS':
            found['reps'] = c
    if 'order' in found and 'exercise' in found:
        found.setdefault('reps', found['exercise'] + 1)
        return found
    return None


def parse_sheet(ws):
    workouts = []
    cols = None
    cur = None        # current workout: list of groups
    cur_group = None

    for r in range(1, ws.max_row + 1):
        hc = header_cols(ws, r)
        if hc:
            cols = hc
            cur = []
            workouts.append(cur)
            cur_group = None
            continue
        if cols is None or cur is None:
            continue
        labelv = cell(ws, r, cols['order'])
        namev = cell(ws, r, cols['exercise'])
        repsv = cell(ws, r, cols['reps'])
        lm = LABEL_RE.match(labelv)
        if not (lm and namev and len(namev) > 2):
            continue
        num = int(lm.group(2))
        if num == 1 or cur_group is None:
            cur_group = []
            cur.append(cur_group)
        cur_group.append({'name': namev, **parse_reps(repsv)})

    result = []
    for wk in workouts:
        groups = [g for g in wk if g]
        if not groups:
            continue
        flat = []
        for gi, g in enumerate(groups):
            n = len(g)
            gtype = 'superset' if n == 2 else 'triset' if n == 3 else 'circuit' if n >= 4 else 'regular'
            label = LETTERS[gi] if gi < len(LETTERS) else f'G{gi}'
            for e in g:
                e.update(group_label=label, group_type=gtype, sets=1)
                flat.append(e)
        result.append(flat)
    return result


def _norm(s):
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9 ]', '', s.lower())).strip()


def dump(sheet):
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, FILE), data_only=True)
    workouts = parse_sheet(wb[sheet])
    print(f"{sheet}: {len(workouts)} workouts")
    for wi, wk in enumerate(workouts[:2], 1):
        print(f"\n  -- Workout {wi} ({len(wk)} ex) --")
        last = None
        for e in wk:
            if e['group_label'] != last:
                print(f"    [{e['group_label']}] {e['group_type'].upper()}")
                last = e['group_label']
            val = f"{e['duration_secs']}s" if e['time_based'] and e['duration_secs'] else (e['reps'] or '-')
            side = f" /{e['per_side']}" if e['per_side'] else ''
            print(f"        - {e['name'][:42]}  ({val}{side})")


def run_import():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    idx = {_norm(r['name']): r['id'] for r in db.execute('SELECT id, name FROM exercises').fetchall()}

    def find_or_create(name):
        k = _norm(name)
        if k in idx:
            return idx[k]
        c = db.execute('INSERT INTO exercises (name) VALUES (?)', (name,))
        idx[k] = c.lastrowid
        return c.lastrowid

    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, FILE), data_only=True)
    for sheet in wb.sheetnames:
        title = f"{PREFIX} | {sheet}"
        prog = db.execute('SELECT id FROM programs WHERE title = ?', (title,)).fetchone()
        if not prog:
            print(f"  SKIP (no program): {title}")
            continue
        pid = prog['id']
        workouts = parse_sheet(wb[sheet])
        if not workouts:
            print(f"  {title}: nothing parsed")
            continue
        old = [r['id'] for r in db.execute('SELECT id FROM workouts WHERE program_id=?', (pid,)).fetchall()]
        for wid in old:
            for x in db.execute('SELECT id FROM workout_exercises WHERE workout_id=?', (wid,)).fetchall():
                db.execute('DELETE FROM workout_exercise_meta WHERE workout_exercise_id=?', (x['id'],))
            db.execute('DELETE FROM workout_exercises WHERE workout_id=?', (wid,))
        db.execute('DELETE FROM workouts WHERE program_id=?', (pid,))

        wpw = 3
        db.execute('UPDATE programs SET duration_weeks=?, workouts_per_week=? WHERE id=?',
                   (max(1, (len(workouts) + wpw - 1) // wpw), wpw, pid))
        ex_total = 0
        for i, wk in enumerate(workouts):
            week, day = i // wpw + 1, i % wpw + 1
            c = db.execute(
                """INSERT INTO workouts (program_id, week_number, day_number, title, duration_mins, intensity, body_parts, workout_type, status)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (pid, week, day, f"{sheet} W{week} - Session {day}", 40, 'Medium', sheet, 'strength', 'draft'))
            wid = c.lastrowid
            for order, e in enumerate(wk):
                ex_id = find_or_create(e['name'])
                c2 = db.execute(
                    """INSERT INTO workout_exercises
                       (workout_id, exercise_id, order_index, sets, reps, duration_secs, rest_secs, group_type, group_label)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (wid, ex_id, order, e['sets'], e['reps'] if e['reps'] is not None else '',
                     e['duration_secs'], 30, e['group_type'], e['group_label']))
                db.execute(
                    """INSERT INTO workout_exercise_meta (workout_exercise_id, per_side, time_based, duration_secs, tracking_type)
                       VALUES (?,?,?,?,?)""",
                    (c2.lastrowid, e['per_side'], e['time_based'], e['duration_secs'], e['tracking_type']))
                ex_total += 1
        db.commit()
        print(f"  {title} (id={pid}): {len(workouts)} workouts, {ex_total} exercises")
    db.execute("DELETE FROM explore_section_items WHERE item_type='workout' AND NOT EXISTS (SELECT 1 FROM workouts w WHERE w.id=explore_section_items.item_id)")
    db.commit()
    db.close()


if __name__ == '__main__':
    if len(sys.argv) >= 3 and sys.argv[1] == '--dump':
        dump(sys.argv[2])
    elif len(sys.argv) >= 2 and sys.argv[1] == '--import':
        run_import()
    else:
        print('usage: parse_bodyweight.py [--dump "<sheet>" | --import]')
