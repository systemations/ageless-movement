#!/usr/bin/env python3
"""Parser for the 2023 PFP membership-program spreadsheet (and same-layout
files: Bodyweight Zone, Buddy Programs, Prehab Advanced).

Each sheet lays out THREE sessions side-by-side in column blocks starting at
columns 2, 8 and 14. Within a block the columns are:
    time/section | label (A1/B1...) | name | sets (xN) | reps

Weeks are stacked vertically, marked by a "W<n>" token in the time column.
Exercises are grouped by their letter (A1/A2 -> group A) so supersets stay
separate, exactly like the AMS import. Sets come from "xN" (inherited across a
group), and reps/duration/per-side are parsed from the reps cell.

  --dump "<sheet>"   print parsed structure for one sheet (no DB writes)
  --import           rebuild all configured sheets into the DB
"""

import os
import re
import sys
import sqlite3

import openpyxl

DB_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'ageless.db')
DATA_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'import-data')

SESSION_BASES = [2, 8, 14]  # 1-indexed column of the time/section cell per session
LABEL_RE = re.compile(r'^([A-Z])(\d+)$')
WEEK_RE = re.compile(r'^W\s*(\d+)$', re.IGNORECASE)
SECTION_WORDS = ('WARM UP', 'WARMUP', 'CORE', 'MOBILITY', 'STRENGTH', 'STRETCH',
                 'CONDITIONING', 'FINISHER', 'COOL', 'PREP', 'PAILS', 'AMRAP', 'SET')

# (file, [sheet names]) -> program title is "<prefix> | <sheet>"
SOURCES = [
    ('2023 - PFP ONLINE MEMBERSHIP PROGRAMS.xlsx', 'PFP', None),  # None = all sheets except skip
]
SKIP_SHEETS = {'Program Templates', 'MEMBERSHIP PROGRAMS'}


def cell_str(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return re.sub(r'\s+', ' ', str(v)).strip()


def is_section(val):
    u = val.upper()
    return any(w in u for w in SECTION_WORDS)


def parse_reps(raw):
    """Return dict(reps, duration_secs, time_based, per_side, tracking_type)."""
    out = {'reps': None, 'duration_secs': None, 'time_based': 0,
           'per_side': None, 'tracking_type': None}
    if not raw:
        return out
    val = raw.strip()
    # per side: "e/s", "e.s", "e s", "each side", "/ side"
    if re.search(r'e\s*[./]\s*s|each side|/\s*side|per side', val, re.IGNORECASE):
        out['per_side'] = 'side'
    val_clean = re.sub(r'\s*e\s*[./]\s*s\.?', '', val, flags=re.IGNORECASE).strip()
    # pure time: "60s", "45s", "1:30", "30-45s"
    mt = re.match(r'^(\d+):(\d{2})$', val_clean)
    ms = re.match(r'^(\d+)\s*s(ec)?$', val_clean, re.IGNORECASE)
    if mt:
        out['duration_secs'] = int(mt.group(1)) * 60 + int(mt.group(2))
        out['time_based'] = 1
        out['tracking_type'] = 'Duration'
        out['reps'] = val_clean
        return out
    if ms:
        out['duration_secs'] = int(ms.group(1))
        out['time_based'] = 1
        out['tracking_type'] = 'Duration'
        out['reps'] = val_clean
        return out
    # reps: "12r", "15-20r", "10" (optionally trailing text kept)
    mr = re.match(r'^(\d+(?:\s*-\s*\d+)?)\s*r\b', val_clean, re.IGNORECASE)
    if mr:
        out['reps'] = re.sub(r'\s*', '', mr.group(1))
        out['tracking_type'] = 'Repetitions'
        return out
    # otherwise keep the cleaned text verbatim (e.g. "1 lap", "10r + 10s")
    out['reps'] = val_clean
    return out


def parse_sheet(ws, title_token=None):
    """Return list of workouts: {week, day, exercises:[{label_letter, name,
    sets, reps fields...}]}.

    Weeks are marked by a "W<n>" token where present; sheets without W# markers
    (e.g. Fundamentals) delimit weeks by the repeating program-title row, so we
    fall back to incrementing the week each time that title reappears."""
    workouts = {}  # (week, day) -> list of exercise dicts (in order)
    group_sets = {}  # (week, day, letter) -> sets (from xN)

    # Does any session column use explicit W# markers?
    has_week_markers = False
    for base in SESSION_BASES:
        for r in range(1, ws.max_row + 1):
            if WEEK_RE.match(cell_str(ws, r, base)):
                has_week_markers = True
                break
        if has_week_markers:
            break

    for day_idx, base in enumerate(SESSION_BASES, start=1):
        c_time, c_label, c_name, c_sets, c_reps = base, base + 1, base + 2, base + 3, base + 4
        week = None
        seen_ex_this_week = False
        for r in range(1, ws.max_row + 1):
            timev = cell_str(ws, r, c_time)
            labelv = cell_str(ws, r, c_label)
            namev = cell_str(ws, r, c_name)
            setsv = cell_str(ws, r, c_sets)
            repsv = cell_str(ws, r, c_reps)

            wm = WEEK_RE.match(timev)
            if wm:
                week = int(wm.group(1))
                seen_ex_this_week = False
                continue
            # title-row week boundary (only when the sheet has no W# markers)
            if not has_week_markers and title_token and title_token in timev.upper():
                week = (week or 0) + 1 if seen_ex_this_week else (week or 1)
                seen_ex_this_week = False
                continue
            lm = LABEL_RE.match(labelv)
            if lm and namev and namev != '-':
                if week is None:
                    week = 1
                letter = lm.group(1)
                key = (week, day_idx)
                wk = workouts.setdefault(key, [])
                sm = re.search(r'(\d+)', setsv)
                if sm and setsv.lower().startswith('x'):
                    group_sets[(week, day_idx, letter)] = int(sm.group(1))
                rp = parse_reps(repsv)
                wk.append({'letter': letter, 'name': namev, **rp})
                seen_ex_this_week = True

    # finalise: assign group_type by letter-group size, sets per group
    result = []
    for (week, day), exs in sorted(workouts.items()):
        # count items per letter
        counts = {}
        for e in exs:
            counts[e['letter']] = counts.get(e['letter'], 0) + 1
        for e in exs:
            n = counts[e['letter']]
            e['group_type'] = ('superset' if n == 2 else 'triset' if n == 3
                               else 'circuit' if n >= 4 else 'regular')
            e['group_label'] = e['letter']
            e['sets'] = group_sets.get((week, day, e['letter']), 1)
        result.append({'week': week, 'day': day, 'exercises': exs})
    return result


def _norm(s):
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9 ]', '', s.lower())).strip()


def title_token_for(sheet_name):
    """First alphabetic word of the sheet name, uppercased — matches the
    repeating title row used as a week boundary on W#-less sheets."""
    m = re.match(r'([A-Za-z]+)', sheet_name)
    return m.group(1).upper() if m else None


def dump(sheet_name):
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, SOURCES[0][0]), data_only=True)
    ws = wb[sheet_name]
    workouts = parse_sheet(ws, title_token_for(sheet_name))
    print(f"{sheet_name}: {len(workouts)} workouts")
    for w in workouts[:3]:
        print(f"\n  -- Week {w['week']} · Session {w['day']} ({len(w['exercises'])} ex) --")
        last = None
        for e in w['exercises']:
            if e['group_label'] != last:
                print(f"    [{e['group_label']}] {e['group_type'].upper()} x{e['sets']}")
                last = e['group_label']
            val = (f"{e['duration_secs']}s" if e['time_based'] and e['duration_secs']
                   else (e['reps'] or '-'))
            side = f" /{e['per_side']}" if e['per_side'] else ''
            print(f"        - {e['name'][:46]}  ({val}{side})")


def run_import():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    idx = {_norm(r['name']): r['id'] for r in db.execute('SELECT id, name FROM exercises').fetchall()}

    def find_or_create_ex(name):
        k = _norm(name)
        if k in idx:
            return idx[k]
        cur = db.execute('INSERT INTO exercises (name) VALUES (?)', (name,))
        idx[k] = cur.lastrowid
        return cur.lastrowid

    fname, prefix, _ = SOURCES[0]
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, fname), data_only=True)
    for sheet in wb.sheetnames:
        if sheet in SKIP_SHEETS:
            continue
        title = f"{prefix} | {sheet}"
        prog = db.execute('SELECT id FROM programs WHERE title = ?', (title,)).fetchone()
        if not prog:
            print(f"  SKIP (no program): {title}")
            continue
        pid = prog['id']
        workouts = parse_sheet(wb[sheet], title_token_for(sheet))
        if not workouts:
            print(f"  {title}: no exercises parsed")
            continue
        # wipe existing workouts + their exercise rows for a clean rebuild
        old_w = [r['id'] for r in db.execute('SELECT id FROM workouts WHERE program_id=?', (pid,)).fetchall()]
        for wid in old_w:
            we = [r['id'] for r in db.execute('SELECT id FROM workout_exercises WHERE workout_id=?', (wid,)).fetchall()]
            for x in we:
                db.execute('DELETE FROM workout_exercise_meta WHERE workout_exercise_id=?', (x,))
            db.execute('DELETE FROM workout_exercises WHERE workout_id=?', (wid,))
        db.execute('DELETE FROM workouts WHERE program_id=?', (pid,))

        n_weeks = max(w['week'] for w in workouts)
        n_days = max(w['day'] for w in workouts)
        db.execute('UPDATE programs SET duration_weeks=?, workouts_per_week=? WHERE id=?',
                   (n_weeks, n_days, pid))
        ex_total = 0
        for w in workouts:
            cur = db.execute(
                """INSERT INTO workouts (program_id, week_number, day_number, title, duration_mins, intensity, body_parts, workout_type, status)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (pid, w['week'], w['day'], f"{sheet} W{w['week']} - Session {w['day']}",
                 40, 'Medium', sheet, 'mobility', 'draft'))
            wid = cur.lastrowid
            for order, e in enumerate(w['exercises']):
                ex_id = find_or_create_ex(e['name'])
                cur2 = db.execute(
                    """INSERT INTO workout_exercises
                       (workout_id, exercise_id, order_index, sets, reps, duration_secs, rest_secs, group_type, group_label)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (wid, ex_id, order, e['sets'], e['reps'] if e['reps'] is not None else '',
                     e['duration_secs'], 30, e['group_type'], e['group_label']))
                we_id = cur2.lastrowid
                db.execute(
                    """INSERT INTO workout_exercise_meta
                       (workout_exercise_id, per_side, time_based, duration_secs, tracking_type)
                       VALUES (?,?,?,?,?)""",
                    (we_id, e['per_side'], e['time_based'], e['duration_secs'], e['tracking_type']))
                ex_total += 1
        db.commit()
        print(f"  {title} (id={pid}): {len(workouts)} workouts, {ex_total} exercises")
    db.close()


if __name__ == '__main__':
    if len(sys.argv) >= 3 and sys.argv[1] == '--dump':
        dump(sys.argv[2])
    elif len(sys.argv) >= 2 and sys.argv[1] == '--import':
        run_import()
    else:
        print('usage: parse_pfp.py [--dump "<sheet>" | --import]')
