#!/usr/bin/env python3
"""Import programs from XLSX files - V2 with proper PFP parsing."""

import openpyxl
import sqlite3
import os
import re

DB_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'ageless.db')
BASE_DIR = os.path.join(os.path.dirname(__file__), '..', '..', '..', '..')

db = sqlite3.connect(DB_PATH)
db.row_factory = sqlite3.Row

exercises_db = {row['name'].lower().strip(): dict(row) for row in db.execute('SELECT * FROM exercises').fetchall()}

def find_exercise(name):
    name = name.lower().strip()
    # Remove common prefixes like "A1 ", "B2 " etc
    name = re.sub(r'^[a-z]\d+\s*', '', name).strip()
    if not name or len(name) < 3:
        return None
    if name in exercises_db:
        return exercises_db[name]
    for key, ex in exercises_db.items():
        if name in key or key in name:
            return ex
    words = [w for w in name.split() if len(w) > 3]
    if words:
        for key, ex in exercises_db.items():
            if all(w in key for w in words):
                return ex
    return None

def get_or_create_exercise(name, body_part=''):
    name = name.strip()
    if not name or len(name) < 3:
        return None
    ex = find_exercise(name)
    if ex:
        return ex['id']
    cur = db.execute('INSERT INTO exercises (name, body_part) VALUES (?, ?)', (name, body_part))
    db.commit()
    exercises_db[name.lower()] = {'id': cur.lastrowid, 'name': name}
    return cur.lastrowid

def get_or_create_program(title, desc='', weeks=4, wpw=3):
    row = db.execute('SELECT id FROM programs WHERE title = ?', (title,)).fetchone()
    if row:
        return row['id']
    cur = db.execute('INSERT INTO programs (coach_id, title, description, duration_weeks, workouts_per_week) VALUES (1, ?, ?, ?, ?)',
                     (title, desc, weeks, wpw))
    db.commit()
    return cur.lastrowid

def get_or_create_workout(prog_id, week, day, title, mins=30, intensity='Medium', parts='', wtype='strength'):
    row = db.execute('SELECT id FROM workouts WHERE program_id=? AND week_number=? AND day_number=?', (prog_id, week, day)).fetchone()
    if row:
        return row['id']
    cur = db.execute('INSERT INTO workouts (program_id, week_number, day_number, title, duration_mins, intensity, body_parts, workout_type) VALUES (?,?,?,?,?,?,?,?)',
                     (prog_id, week, day, title, mins, intensity, parts, wtype))
    db.commit()
    return cur.lastrowid

def add_we(w_id, ex_id, order, sets=1, reps='10', group=None):
    exists = db.execute('SELECT id FROM workout_exercises WHERE workout_id=? AND exercise_id=? AND order_index=?', (w_id, ex_id, order)).fetchone()
    if exists:
        return
    db.execute('INSERT INTO workout_exercises (workout_id, exercise_id, order_index, sets, reps, group_type) VALUES (?,?,?,?,?,?)',
               (w_id, ex_id, order, sets, reps, group))
    db.commit()

def is_exercise_label(val):
    """Check if value looks like A1, B2, C3, etc."""
    return bool(re.match(r'^[A-Z]\d+$', val.strip()))

def is_section_header(val):
    """Check if value is a section header."""
    headers = ['WARM UP', 'WARMUP', 'SET', 'AMRAP', 'FINISHER', 'STRETCH', 'COOL DOWN',
               'MOBILITY', 'STRENGTH', 'PAILS', 'PREP', 'CONDITIONING']
    return any(h in val.upper() for h in headers)

def parse_pfp_sheet(wb, sheet_name, program_title=None):
    """Parse a PFP program sheet with A1/B1/C1 exercise format."""
    ws = wb[sheet_name]
    prog_title = program_title or f'PFP | {sheet_name}'

    # Determine type
    wtype = 'mobility'
    sn = sheet_name.lower()
    if any(k in sn for k in ['conditioning', 'strength', 'rings', 'handstand', 'tumbling']):
        wtype = 'strength'
    elif 'stretch' in sn:
        wtype = 'flexibility'
    elif 'prehab' in sn:
        wtype = 'rehab'
    elif 'kinstretch' in sn:
        wtype = 'mobility'

    prog_id = get_or_create_program(prog_title, f'{sheet_name} program', weeks=8, wpw=2)

    current_week = 0
    workout_id = None
    exercise_order = 0
    group_type = None
    workout_count = 0
    exercise_count = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        vals = [str(v).strip() if v else '' for v in row[:8]]

        # Detect week (W1, W2, etc.)
        for v in vals:
            m = re.match(r'W\s*(\d+)', v)
            if m:
                current_week = int(m.group(1))

        # Detect section headers -> create workout
        if any(is_section_header(v) for v in vals if v):
            header = next((v for v in vals if v and is_section_header(v)), '')
            if 'WARM' in header.upper():
                group_type = 'warmup'
                workout_count += 1
                week = current_week if current_week > 0 else (workout_count - 1) // 2 + 1
                day = (workout_count - 1) % 2 + 1
                focus = vals[6] if len(vals) > 6 and vals[6] else ''
                workout_id = get_or_create_workout(prog_id, week, day,
                    f'{sheet_name} W{week} - Session {day}' + (f' ({focus[:30]})' if focus else ''),
                    mins=45, intensity='Medium', parts=sheet_name, wtype=wtype)
                exercise_order = 0
            elif 'SET' in header.upper() or 'AMRAP' in header.upper():
                group_type = 'superset'
            elif 'FINISHER' in header.upper():
                group_type = 'superset'
            elif 'STRETCH' in header.upper() or 'COOL' in header.upper():
                group_type = None
            elif 'MOBILITY' in header.upper():
                group_type = None
            elif 'STRENGTH' in header.upper():
                group_type = 'superset'
            elif 'PAILS' in header.upper():
                group_type = None
            continue

        if not workout_id:
            continue

        # Detect exercises - look for A1/B1 pattern in any column
        for i in range(len(vals) - 1):
            if vals[i] and is_exercise_label(vals[i]):
                # Exercise name is in next column
                ex_name = vals[i + 1] if i + 1 < len(vals) else ''
                if not ex_name:
                    # Try column after
                    ex_name = vals[i + 2] if i + 2 < len(vals) else ''
                if ex_name and len(ex_name) > 3:
                    # Get reps from remaining columns
                    reps = '10'
                    sets = 1
                    for j in range(i + 2, min(i + 5, len(vals))):
                        v = vals[j]
                        if v:
                            if re.match(r'x\d', v, re.IGNORECASE):
                                try:
                                    sets = int(re.search(r'\d+', v).group())
                                except:
                                    pass
                            elif any(c in v for c in ['r', 's', 'min', 'sec', 'rep']):
                                reps = v
                            elif not reps or reps == '10':
                                reps = v

                    ex_id = get_or_create_exercise(ex_name, sheet_name)
                    if ex_id:
                        add_we(workout_id, ex_id, exercise_order, sets, reps, group_type)
                        exercise_order += 1
                        exercise_count += 1
                break

        # Also check for exercises without A1/B1 labels (some sheets just list names)
        if not any(is_exercise_label(v) for v in vals):
            for i, v in enumerate(vals):
                if v and len(v) > 5 and not v.startswith('W') and not is_section_header(v) and not v.isdigit():
                    if not any(v.startswith(x) for x in ['Done', '✔', 'TUE', 'MON', 'WED', 'THU', 'FRI', 'SAT', 'SUN', '$']):
                        ex = find_exercise(v)
                        if ex:
                            reps = vals[i + 1] if i + 1 < len(vals) and vals[i + 1] else '10'
                            add_we(workout_id, ex['id'], exercise_order, 1, reps, group_type)
                            exercise_order += 1
                            exercise_count += 1
                            break

    return workout_count, exercise_count

# ===== MAIN =====
print(f"Exercises in DB: {len(exercises_db)}")

# Parse PFP Online
filepath = os.path.join(BASE_DIR, '2023 - PFP ONLINE MEMBERSHIP PROGRAMS.xlsx')
if os.path.exists(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    skip = ['Program Templates', 'MEMBERSHIP PROGRAMS']
    for sn in wb.sheetnames:
        if sn in skip:
            continue
        wc, ec = parse_pfp_sheet(wb, sn)
        print(f"  PFP | {sn}: {wc} workouts, {ec} exercises linked")

# Parse Bodyweight Zone
filepath = os.path.join(BASE_DIR, 'Bodyweight Zone.xlsx')
if os.path.exists(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for sn in wb.sheetnames:
        wc, ec = parse_pfp_sheet(wb, sn, f'Bodyweight | {sn}')
        print(f"  Bodyweight | {sn}: {wc} workouts, {ec} exercises linked")

# Parse Buddy Programs
filepath = os.path.join(BASE_DIR, 'Buddy Programs.xlsx')
if os.path.exists(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for sn in wb.sheetnames:
        wc, ec = parse_pfp_sheet(wb, sn, f'Buddy | {sn}')
        print(f"  Buddy | {sn}: {wc} workouts, {ec} exercises linked")

# Parse Prehab Advanced
filepath = os.path.join(BASE_DIR, 'prehab for performance-adv.xlsx')
if os.path.exists(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for sn in wb.sheetnames:
        wc, ec = parse_pfp_sheet(wb, sn, f'PFP Advanced | {sn}')
        print(f"  PFP Adv | {sn}: {wc} workouts, {ec} exercises linked")

# Summary
pc = db.execute('SELECT COUNT(*) as c FROM programs').fetchone()['c']
wc = db.execute('SELECT COUNT(*) as c FROM workouts').fetchone()['c']
ec = db.execute('SELECT COUNT(*) as c FROM exercises').fetchone()['c']
lc = db.execute('SELECT COUNT(*) as c FROM workout_exercises').fetchone()['c']
print(f"\n=== TOTAL ===")
print(f"  Programs: {pc}")
print(f"  Workouts: {wc}")
print(f"  Exercises: {ec}")
print(f"  Exercise links: {lc}")
db.close()
