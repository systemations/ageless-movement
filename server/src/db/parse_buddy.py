#!/usr/bin/env python3
"""Parser for Buddy Programs.xlsx.

Single column block (cols 2-6): TIME | ORDER | EXERCISE | REPS | SETS.
Workouts are stacked vertically, each starting with a "WORKOUT OF THE DAY"
banner. The ORDER column holds A1/A2/B1... labels (grouping key), plus
"Warm Up" / "Workout (n)" for warm-ups and AMRAP blocks. Letter grouping works
exactly like the PFP/AMS imports.

  --dump "<sheet>"   print parsed structure (no DB writes)
  --import           rebuild all Buddy programs
"""

import os
import re
import sys
import sqlite3

import openpyxl

DB_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'ageless.db')
DATA_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'import-data')
FILE = 'Buddy Programs.xlsx'
PREFIX = 'Buddy'

C_TIME, C_LABEL, C_NAME, C_REPS, C_SETS = 2, 3, 4, 5, 6
LABEL_RE = re.compile(r'^([A-Z])(\d+)$')
WORKOUT_BANNER = 'WORKOUT OF THE DAY'


def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return '' if v is None else re.sub(r'\s+', ' ', str(v)).strip()


def parse_reps(raw):
    out = {'reps': None, 'duration_secs': None, 'time_based': 0, 'per_side': None, 'tracking_type': None}
    if not raw:
        return out
    val = raw.strip()
    if re.search(r'e\s*[./]\s*s|each side|per side', val, re.IGNORECASE):
        out['per_side'] = 'side'
    elif re.search(r'each direction', val, re.IGNORECASE):
        out['per_side'] = 'direction'
    clean = re.sub(r'\s*(e\s*[./]\s*s\.?|each side|each direction|per side)', '', val, flags=re.IGNORECASE).strip()
    ms = re.match(r'^(\d+)\s*-?\s*(\d+)?\s*s(ec)?$', clean, re.IGNORECASE)
    mt = re.match(r'^(\d+):(\d{2})$', clean)
    if mt:
        out.update(duration_secs=int(mt.group(1)) * 60 + int(mt.group(2)), time_based=1, tracking_type='Duration', reps=clean)
        return out
    if ms and 's' in clean.lower() and not re.search(r'r$', clean):
        out.update(duration_secs=int(ms.group(1)), time_based=1, tracking_type='Duration', reps=clean)
        return out
    mr = re.match(r'^(\d+(?:\.\d+)?(?:\s*-\s*\d+)?)\s*(r(eps)?)?\b', clean, re.IGNORECASE)
    if mr and mr.group(1):
        num = re.sub(r'\s*', '', mr.group(1))
        num = re.sub(r'\.0$', '', num)
        out['reps'] = num
        out['tracking_type'] = 'Repetitions'
        return out
    out['reps'] = clean
    return out


def parse_sheet(ws):
    """Return list of workouts (each: list of exercise dicts), in document order."""
    workouts = []
    cur = None  # current workout list
    group_sets = {}  # id(workout)+letter -> sets
    section = ''

    for r in range(1, ws.max_row + 1):
        timev = cell(ws, r, C_TIME)
        labelv = cell(ws, r, C_LABEL)
        namev = cell(ws, r, C_NAME)
        repsv = cell(ws, r, C_REPS)
        setsv = cell(ws, r, C_SETS)

        bu = timev.upper()
        if WORKOUT_BANNER in bu or bu.endswith('WORKOUT'):  # "[..] WORKOUT OF THE DAY" / "MOBILITY WORKOUT"
            cur = []
            workouts.append(cur)
            section = ''
            continue
        # column-header row: TIME | ORDER | EXERCISE | REPS | SETS
        if labelv.upper() == 'ORDER' or namev.upper() == 'EXERCISE' or timev.upper() == 'TIME':
            continue
        # section header: text in time col, nothing in label/name (e.g. "WARM UP", "STRENGTH", "CORE")
        if timev and not labelv and not namev and not re.match(r'^[\d.]+$', timev):
            section = timev.upper()
            continue
        if cur is None or not namev or len(namev) < 3:
            continue

        lm = LABEL_RE.match(labelv)
        is_warm = labelv.lower().startswith('warm') or 'WARM' in section
        if lm:
            letter = lm.group(1)
        elif labelv.lower().startswith('warm'):
            letter = 'WU'
        else:
            # AMRAP / unlabeled named row -> standalone group (no letter)
            letter = None

        sm = re.search(r'(\d+)', setsv)
        if sm and (setsv.lower().startswith('x') or 'round' in setsv.lower()):
            group_sets[(id(cur), letter)] = int(sm.group(1))

        rp = parse_reps(repsv)
        cur.append({'letter': letter, 'name': namev, 'is_warm': is_warm, **rp})

    # finalise grouping
    result = []
    for wk in workouts:
        if not wk:
            continue
        counts = {}
        for e in wk:
            if e['letter']:
                counts[e['letter']] = counts.get(e['letter'], 0) + 1
        for e in wk:
            n = counts.get(e['letter'], 1)
            if e['is_warm']:
                e['group_type'] = 'warmup'
            else:
                e['group_type'] = ('superset' if n == 2 else 'triset' if n == 3
                                   else 'circuit' if n >= 4 else 'regular')
            e['group_label'] = e['letter']  # 'A'/'B'/'WU' or None for standalone
            e['sets'] = group_sets.get((id(wk), e['letter']), 1)
        result.append(wk)
    return result


def _norm(s):
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9 ]', '', s.lower())).strip()


def dump(sheet):
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, FILE), data_only=True)
    workouts = parse_sheet(wb[sheet])
    print(f"{sheet}: {len(workouts)} workouts")
    for wi, wk in enumerate(workouts[:2], 1):
        print(f"\n  -- Workout {wi} ({len(wk)} ex) --")
        last = None
        for e in wk:
            if e['group_label'] != last:
                print(f"    [{e['group_label']}] {e['group_type'].upper()} x{e['sets']}")
                last = e['group_label']
            val = f"{e['duration_secs']}s" if e['time_based'] and e['duration_secs'] else (e['reps'] or '-')
            side = f" /{e['per_side']}" if e['per_side'] else ''
            print(f"        - {e['name'][:44]}  ({val}{side})")


def run_import():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    idx = {_norm(r['name']): r['id'] for r in db.execute('SELECT id, name FROM exercises').fetchall()}

    def find_or_create(name):
        k = _norm(name)
        if k in idx:
            return idx[k]
        c = db.execute('INSERT INTO exercises (name) VALUES (?)', (name,))
        idx[k] = c.lastrowid
        return c.lastrowid

    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, FILE), data_only=True)
    for sheet in wb.sheetnames:
        title = f"{PREFIX} | {sheet}"
        prog = db.execute('SELECT id FROM programs WHERE title = ?', (title,)).fetchone()
        if not prog:
            print(f"  SKIP (no program): {title}")
            continue
        pid = prog['id']
        workouts = parse_sheet(wb[sheet])
        if not workouts:
            print(f"  {title}: nothing parsed")
            continue
        old = [r['id'] for r in db.execute('SELECT id FROM workouts WHERE program_id=?', (pid,)).fetchall()]
        for wid in old:
            for x in db.execute('SELECT id FROM workout_exercises WHERE workout_id=?', (wid,)).fetchall():
                db.execute('DELETE FROM workout_exercise_meta WHERE workout_exercise_id=?', (x['id'],))
            db.execute('DELETE FROM workout_exercises WHERE workout_id=?', (wid,))
        db.execute('DELETE FROM workouts WHERE program_id=?', (pid,))

        wpw = 2  # Buddy programs run twice a week
        db.execute('UPDATE programs SET duration_weeks=?, workouts_per_week=? WHERE id=?',
                   (max(1, (len(workouts) + wpw - 1) // wpw), wpw, pid))
        ex_total = 0
        for i, wk in enumerate(workouts):
            week = i // wpw + 1
            day = i % wpw + 1
            cur = db.execute(
                """INSERT INTO workouts (program_id, week_number, day_number, title, duration_mins, intensity, body_parts, workout_type, status)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (pid, week, day, f"{sheet} W{week} - Session {day}", 40, 'Medium', sheet, 'strength', 'draft'))
            wid = cur.lastrowid
            for order, e in enumerate(wk):
                ex_id = find_or_create(e['name'])
                c2 = db.execute(
                    """INSERT INTO workout_exercises
                       (workout_id, exercise_id, order_index, sets, reps, duration_secs, rest_secs, group_type, group_label)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (wid, ex_id, order, e['sets'], e['reps'] if e['reps'] is not None else '',
                     e['duration_secs'], 30, e['group_type'], e['group_label']))
                db.execute(
                    """INSERT INTO workout_exercise_meta (workout_exercise_id, per_side, time_based, duration_secs, tracking_type)
                       VALUES (?,?,?,?,?)""",
                    (c2.lastrowid, e['per_side'], e['time_based'], e['duration_secs'], e['tracking_type']))
                ex_total += 1
        db.commit()
        print(f"  {title} (id={pid}): {len(workouts)} workouts, {ex_total} exercises")
    # clean orphaned explore items pointing at deleted workouts
    db.execute("DELETE FROM explore_section_items WHERE item_type='workout' AND NOT EXISTS (SELECT 1 FROM workouts w WHERE w.id=explore_section_items.item_id)")
    db.commit()
    db.close()


if __name__ == '__main__':
    if len(sys.argv) >= 3 and sys.argv[1] == '--dump':
        dump(sys.argv[2])
    elif len(sys.argv) >= 2 and sys.argv[1] == '--import':
        run_import()
    else:
        print('usage: parse_buddy.py [--dump "<sheet>" | --import]')
