#!/usr/bin/env python3
"""Parser for the PFP "Stretch Lower" / "Stretch Upper" sheets.

These are sequential stretch routines (NO A1/B1 supersets), laid out in two
session blocks. Each block has a "Warm Up" section (stretch names in the order
column) and a "Main Program" section with an "Exercise | Action | Time" header.
We emit each session as a workout of straight-set / warm-up stretches.

  --dump "<sheet>" | --import
"""

import os
import re
import sys
import sqlite3

import openpyxl

DB_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'ageless.db')
DATA_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'import-data')
FILE = '2023 - PFP ONLINE MEMBERSHIP PROGRAMS.xlsx'
PREFIX = 'PFP'
SHEETS = ['Stretch Lower', 'Stretch Upper']

# (order/warmup-name col, main-program name col, time col)
BLOCKS = [(2, 3, 5), (9, 10, 12)]


def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return '' if v is None else re.sub(r'\s+', ' ', str(v)).strip()


def parse_time(raw):
    out = {'reps': raw or None, 'duration_secs': None, 'time_based': 0, 'per_side': None, 'tracking_type': None}
    if not raw:
        return out
    if re.search(r'e\s*[./]\s*s|each side', raw, re.IGNORECASE):
        out['per_side'] = 'side'
    m = re.search(r'(\d+)\s*-\s*(\d+)\s*s|(\d+)\s*s\b', raw)
    if m and ('s' in raw.lower()):
        out['time_based'] = 1
        out['tracking_type'] = 'Duration'
    return out


def parse_sheet(ws):
    """Return up to 2 workouts (one per session block)."""
    workouts = []
    for name_w, name_m, time_c in BLOCKS:
        exs = []
        section = ''
        seen_main = False
        for r in range(1, ws.max_row + 1):
            head = cell(ws, r, name_w)
            hu = head.upper()
            if hu.startswith('WARM UP'):
                if seen_main:
                    break  # next week's repetition of the same routine - stop
                section = 'warmup'
                continue
            if hu.startswith('MAIN PROGRAM'):
                section = 'main'
                continue
            if cell(ws, r, name_m).upper() == 'EXERCISE':  # column header row
                continue
            if section == 'warmup':
                nm = head
                if not nm or re.match(r'^[\d.]+$', nm) or '|' in nm or nm.upper().startswith('W ') or len(nm) < 4:
                    continue
                t = parse_time(cell(ws, r, time_c))
                exs.append({'name': nm, 'warm': True, **t})
            elif section == 'main':
                nm = cell(ws, r, name_m)
                if not nm or len(nm) < 4 or nm.upper() == 'ACTION':
                    continue
                t = parse_time(cell(ws, r, time_c))
                exs.append({'name': nm, 'warm': False, **t})
                seen_main = True
        if exs:
            workouts.append(exs)
    return workouts


def _norm(s):
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9 ]', '', s.lower())).strip()


def dump(sheet):
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, FILE), data_only=True)
    workouts = parse_sheet(wb[sheet])
    print(f"{sheet}: {len(workouts)} sessions")
    for wi, wk in enumerate(workouts, 1):
        print(f"\n  -- Session {wi} ({len(wk)} stretches) --")
        for e in wk:
            tag = 'warmup' if e['warm'] else 'stretch'
            print(f"    [{tag}] {e['name'][:40]}  ({e['reps'] or '-'})")


def run_import():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    idx = {_norm(r['name']): r['id'] for r in db.execute('SELECT id, name FROM exercises').fetchall()}

    def find_or_create(name):
        k = _norm(name)
        if k in idx:
            return idx[k]
        c = db.execute('INSERT INTO exercises (name) VALUES (?)', (name,))
        idx[k] = c.lastrowid
        return c.lastrowid

    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, FILE), data_only=True)
    for sheet in SHEETS:
        title = f"{PREFIX} | {sheet}"
        prog = db.execute('SELECT id FROM programs WHERE title = ?', (title,)).fetchone()
        if not prog:
            print(f"  SKIP (no program): {title}")
            continue
        pid = prog['id']
        workouts = parse_sheet(wb[sheet])
        if not workouts:
            print(f"  {title}: nothing parsed")
            continue
        for wid in [r['id'] for r in db.execute('SELECT id FROM workouts WHERE program_id=?', (pid,)).fetchall()]:
            for x in db.execute('SELECT id FROM workout_exercises WHERE workout_id=?', (wid,)).fetchall():
                db.execute('DELETE FROM workout_exercise_meta WHERE workout_exercise_id=?', (x['id'],))
            db.execute('DELETE FROM workout_exercises WHERE workout_id=?', (wid,))
        db.execute('DELETE FROM workouts WHERE program_id=?', (pid,))
        db.execute('UPDATE programs SET duration_weeks=?, workouts_per_week=? WHERE id=?', (6, len(workouts), pid))
        ex_total = 0
        for i, wk in enumerate(workouts):
            c = db.execute(
                """INSERT INTO workouts (program_id, week_number, day_number, title, duration_mins, intensity, body_parts, workout_type, status)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (pid, 1, i + 1, f"{sheet} - Session {i + 1}", 30, 'Low', sheet, 'flexibility', 'draft'))
            wid = c.lastrowid
            for order, e in enumerate(wk):
                ex_id = find_or_create(e['name'])
                c2 = db.execute(
                    """INSERT INTO workout_exercises
                       (workout_id, exercise_id, order_index, sets, reps, duration_secs, rest_secs, group_type, group_label)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (wid, ex_id, order, 1, e['reps'] if e['reps'] is not None else '',
                     e['duration_secs'], 30, 'warmup' if e['warm'] else 'regular', None))
                db.execute(
                    """INSERT INTO workout_exercise_meta (workout_exercise_id, per_side, time_based, duration_secs, tracking_type)
                       VALUES (?,?,?,?,?)""",
                    (c2.lastrowid, e['per_side'], e['time_based'], e['duration_secs'], e['tracking_type']))
                ex_total += 1
        db.commit()
        print(f"  {title} (id={pid}): {len(workouts)} sessions, {ex_total} stretches")
    db.execute("DELETE FROM explore_section_items WHERE item_type='workout' AND NOT EXISTS (SELECT 1 FROM workouts w WHERE w.id=explore_section_items.item_id)")
    db.commit()
    db.close()


if __name__ == '__main__':
    if len(sys.argv) >= 3 and sys.argv[1] == '--dump':
        dump(sys.argv[2])
    elif len(sys.argv) >= 2 and sys.argv[1] == '--import':
        run_import()
    else:
        print('usage: parse_stretch.py [--dump "<sheet>" | --import]')
