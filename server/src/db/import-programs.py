#!/usr/bin/env python3
"""
Import programs from XLSX files into the Ageless Movement database.
Parses the mobility follow-along program structure and creates programs/workouts/exercise links.
"""

import openpyxl
import sqlite3
import os
import re
import sys

DB_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'ageless.db')
BASE_DIR = os.path.join(os.path.dirname(__file__), '..', '..', '..', '..')

db = sqlite3.connect(DB_PATH)
db.row_factory = sqlite3.Row

# Get existing exercises for matching
exercises = {row['name'].lower().strip(): dict(row) for row in db.execute('SELECT * FROM exercises').fetchall()}
print(f"Existing exercises in DB: {len(exercises)}")

def find_exercise(name):
    """Find an exercise by name, fuzzy matching."""
    name_clean = name.lower().strip()
    # Exact match
    if name_clean in exercises:
        return exercises[name_clean]
    # Partial match
    for key, ex in exercises.items():
        if name_clean in key or key in name_clean:
            return ex
    # Word match
    words = name_clean.split()
    for key, ex in exercises.items():
        if all(w in key for w in words if len(w) > 3):
            return ex
    return None

def create_program(title, description='', weeks=4, workouts_per_week=6, coach_id=1):
    """Create a program and return its ID."""
    existing = db.execute('SELECT id FROM programs WHERE title = ?', (title,)).fetchone()
    if existing:
        print(f"  Program '{title}' already exists (id={existing['id']}), skipping")
        return existing['id']

    cur = db.execute(
        'INSERT INTO programs (coach_id, title, description, duration_weeks, workouts_per_week) VALUES (?, ?, ?, ?, ?)',
        (coach_id, title, description, weeks, workouts_per_week)
    )
    db.commit()
    print(f"  Created program: {title} (id={cur.lastrowid})")
    return cur.lastrowid

def create_workout(program_id, week, day, title, duration_mins=20, intensity='Low', body_parts='', workout_type='mobility'):
    """Create a workout and return its ID."""
    existing = db.execute('SELECT id FROM workouts WHERE program_id = ? AND week_number = ? AND day_number = ?',
                         (program_id, week, day)).fetchone()
    if existing:
        return existing['id']

    cur = db.execute(
        'INSERT INTO workouts (program_id, week_number, day_number, title, duration_mins, intensity, body_parts, workout_type) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
        (program_id, week, day, title, duration_mins, intensity, body_parts, workout_type)
    )
    db.commit()
    return cur.lastrowid

def add_exercise_to_workout(workout_id, exercise_id, order_idx, sets=1, reps='10', group_type=None):
    """Link an exercise to a workout."""
    existing = db.execute('SELECT id FROM workout_exercises WHERE workout_id = ? AND exercise_id = ? AND order_index = ?',
                         (workout_id, exercise_id, order_idx)).fetchone()
    if existing:
        return
    db.execute(
        'INSERT INTO workout_exercises (workout_id, exercise_id, order_index, sets, reps, group_type) VALUES (?, ?, ?, ?, ?, ?)',
        (workout_id, exercise_id, order_idx, sets, reps, group_type)
    )

def create_exercise_if_missing(name, body_part='', equipment=''):
    """Create exercise if it doesn't exist."""
    name_clean = name.strip()
    if not name_clean:
        return None
    existing = find_exercise(name_clean)
    if existing:
        return existing['id']

    cur = db.execute(
        'INSERT INTO exercises (name, body_part, equipment) VALUES (?, ?, ?)',
        (name_clean, body_part, equipment)
    )
    db.commit()
    exercises[name_clean.lower()] = {'id': cur.lastrowid, 'name': name_clean}
    return cur.lastrowid

# ===== PARSE MOBILITY FOLLOW ALONG =====
def parse_mobility_follow_along():
    print("\n=== Parsing PFP ONLINE - MOBILITY ===")
    filepath = os.path.join(BASE_DIR, 'PFP ONLINE - MOBILITY (1).xlsx')
    if not os.path.exists(filepath):
        print("  File not found, skipping")
        return

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['PFP | MOBILITY FOLLOW ALONG PRO']

    program_id = create_program('AMS | Mobility Follow Along',
        'Follow-along mobility routines designed to unlock pain-free movement. Each day targets specific body areas.',
        weeks=4, workouts_per_week=6)

    current_day = 0
    current_title = ''
    workout_id = None
    exercise_order = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        vals = [str(v).strip() if v else '' for v in row[:8]]

        # Detect DAY headers
        if vals[1] and vals[1].startswith('DAY'):
            day_match = re.match(r'DAY\s+(\d+)', vals[1])
            if day_match:
                current_day = int(day_match.group(1))
                current_title = vals[3] if vals[3] else f'Day {current_day}'

                # Determine body parts from title
                body_parts = ''
                title_lower = current_title.lower()
                if 'hip' in title_lower: body_parts = 'Hips'
                elif 'spine' in title_lower: body_parts = 'Spine'
                elif 'shoulder' in title_lower: body_parts = 'Shoulders'
                elif 'ankle' in title_lower or 'foot' in title_lower: body_parts = 'Ankles'

                week = (current_day - 1) // 6 + 1
                day_in_week = (current_day - 1) % 6 + 1

                workout_id = create_workout(program_id, week, day_in_week,
                    f'{current_day}. {current_title}',
                    duration_mins=20, intensity='Low', body_parts=body_parts, workout_type='mobility')
                exercise_order = 0
                continue

        # Detect exercise rows (have a number in col B and exercise name in col D)
        if vals[1] and workout_id:
            try:
                ex_num = float(vals[1])
                ex_name = vals[3]
                if ex_name and not ex_name.startswith('$') and not ex_name.startswith('LEARNING'):
                    reps = vals[4] if vals[4] else '10'
                    sets = int(float(vals[5])) if vals[5] and vals[5] not in ['', '0'] else 1

                    ex_id = create_exercise_if_missing(ex_name)
                    if ex_id:
                        add_exercise_to_workout(workout_id, ex_id, exercise_order, sets, reps)
                        exercise_order += 1
            except (ValueError, IndexError):
                pass

    db.commit()
    print(f"  Mobility Follow Along: {current_day} days imported")

# ===== PARSE PFP ONLINE MEMBERSHIP PROGRAMS =====
def parse_pfp_programs():
    print("\n=== Parsing PFP Online Membership Programs ===")
    filepath = os.path.join(BASE_DIR, '2023 - PFP ONLINE MEMBERSHIP PROGRAMS.xlsx')
    if not os.path.exists(filepath):
        print("  File not found, skipping")
        return

    wb = openpyxl.load_workbook(filepath, data_only=True)

    skip_tabs = ['Program Templates', 'MEMBERSHIP PROGRAMS']

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_tabs:
            continue

        ws = wb[sheet_name]
        print(f"\n  --- {sheet_name} ---")

        # Determine workout type
        wtype = 'mobility'
        if 'strength' in sheet_name.lower() or 'conditioning' in sheet_name.lower():
            wtype = 'strength'
        elif 'stretch' in sheet_name.lower():
            wtype = 'flexibility'
        elif 'prehab' in sheet_name.lower():
            wtype = 'rehab'
        elif 'handstand' in sheet_name.lower() or 'rings' in sheet_name.lower():
            wtype = 'strength'

        program_id = create_program(f'PFP | {sheet_name}',
            f'PFP Online {sheet_name} program', weeks=4, workouts_per_week=3)

        current_week = 0
        workout_id = None
        exercise_order = 0
        day_counter = 0

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            vals = [str(v).strip() if v else '' for v in row[:8]]

            # Detect week headers
            for i, v in enumerate(vals):
                if v and re.match(r'W\s*\d+|WEEK|Week', v, re.IGNORECASE):
                    week_match = re.search(r'(\d+)', v)
                    if week_match:
                        current_week = int(week_match.group(1))

            # Detect workout sections (PREP, numbered exercises)
            if vals[1] and vals[1].upper() in ['PREP', 'A', 'B', 'C']:
                day_counter += 1
                week = current_week if current_week > 0 else (day_counter - 1) // 3 + 1
                day_in_week = (day_counter - 1) % 3 + 1

                workout_id = create_workout(program_id, week, day_in_week,
                    f'{sheet_name} - W{week}D{day_in_week}',
                    duration_mins=30, intensity='Medium', body_parts=sheet_name, workout_type=wtype)
                exercise_order = 0
                continue

            # Detect exercise rows
            if workout_id and vals[1]:
                try:
                    ex_num = float(vals[1])
                    # Exercise name could be in different columns
                    ex_name = ''
                    for col in [3, 2, 4]:
                        if col < len(vals) and vals[col] and len(vals[col]) > 3 and not vals[col].startswith('W'):
                            ex_name = vals[col]
                            break

                    if ex_name:
                        reps = ''
                        for col in [4, 5]:
                            if col < len(vals) and vals[col]:
                                reps = vals[col]
                                break
                        if not reps:
                            reps = '10'

                        sets = 1
                        for col in [5, 6]:
                            if col < len(vals) and vals[col]:
                                try:
                                    sets = int(float(vals[col]))
                                except:
                                    pass
                                break

                        ex_id = create_exercise_if_missing(ex_name)
                        if ex_id:
                            add_exercise_to_workout(workout_id, ex_id, exercise_order, sets, reps)
                            exercise_order += 1
                except (ValueError, IndexError):
                    pass

        db.commit()
        print(f"  {sheet_name}: {day_counter} workouts imported")

# ===== PARSE BODYWEIGHT ZONE =====
def parse_bodyweight_zone():
    print("\n=== Parsing Bodyweight Zone ===")
    filepath = os.path.join(BASE_DIR, 'Bodyweight Zone.xlsx')
    if not os.path.exists(filepath):
        print("  File not found, skipping")
        return

    wb = openpyxl.load_workbook(filepath, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  --- {sheet_name} ---")

        program_id = create_program(f'Bodyweight | {sheet_name}',
            f'Bodyweight Zone {sheet_name} program', weeks=4, workouts_per_week=3)

        workout_id = None
        exercise_order = 0
        day_counter = 0

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            vals = [str(v).strip() if v else '' for v in row[:8]]

            # Detect PREP/workout sections
            if any(v.upper() in ['PREP', 'A', 'B', 'C', 'WARM UP', 'WARMUP'] for v in vals if v):
                day_counter += 1
                week = (day_counter - 1) // 3 + 1
                day_in_week = (day_counter - 1) % 3 + 1
                workout_id = create_workout(program_id, week, day_in_week,
                    f'{sheet_name} - Session {day_counter}',
                    duration_mins=45, intensity='Medium', body_parts='Full Body', workout_type='strength')
                exercise_order = 0
                continue

            # Exercise rows
            if workout_id:
                for v in vals:
                    if v and len(v) > 5 and not v.startswith('WEEK') and not v.startswith('Focus') and not v.isdigit():
                        # Could be exercise name
                        if any(c.isalpha() for c in v) and not v.startswith('$'):
                            ex = find_exercise(v)
                            if ex:
                                add_exercise_to_workout(workout_id, ex['id'], exercise_order, 3, '10')
                                exercise_order += 1
                                break

        db.commit()
        print(f"  {sheet_name}: {day_counter} workouts")

# ===== PARSE BUDDY PROGRAMS =====
def parse_buddy_programs():
    print("\n=== Parsing Buddy Programs ===")
    filepath = os.path.join(BASE_DIR, 'Buddy Programs.xlsx')
    if not os.path.exists(filepath):
        print("  File not found, skipping")
        return

    wb = openpyxl.load_workbook(filepath, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  --- {sheet_name} ---")

        program_id = create_program(f'Buddy | {sheet_name}',
            f'Buddy workout program: {sheet_name}', weeks=4, workouts_per_week=2)

        day_counter = 0
        workout_id = None
        exercise_order = 0

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            vals = [str(v).strip() if v else '' for v in row[:8]]

            if any(v.upper() in ['PREP', 'A', 'B', 'WARM UP'] for v in vals if v):
                day_counter += 1
                week = (day_counter - 1) // 2 + 1
                day_in_week = (day_counter - 1) % 2 + 1
                workout_id = create_workout(program_id, week, day_in_week,
                    f'{sheet_name} - Session {day_counter}',
                    duration_mins=45, intensity='Medium', body_parts='Full Body', workout_type='strength')
                exercise_order = 0
                continue

            if workout_id:
                for v in vals:
                    if v and len(v) > 5:
                        ex = find_exercise(v)
                        if ex:
                            add_exercise_to_workout(workout_id, ex['id'], exercise_order, 3, '10')
                            exercise_order += 1
                            break

        db.commit()
        print(f"  {sheet_name}: {day_counter} workouts")

# ===== RUN ALL =====
if __name__ == '__main__':
    print("Starting program import...")
    parse_mobility_follow_along()
    parse_pfp_programs()
    parse_bodyweight_zone()
    parse_buddy_programs()

    # Summary
    program_count = db.execute('SELECT COUNT(*) as c FROM programs').fetchone()['c']
    workout_count = db.execute('SELECT COUNT(*) as c FROM workouts').fetchone()['c']
    exercise_count = db.execute('SELECT COUNT(*) as c FROM exercises').fetchone()['c']
    link_count = db.execute('SELECT COUNT(*) as c FROM workout_exercises').fetchone()['c']

    print(f"\n=== IMPORT COMPLETE ===")
    print(f"  Programs: {program_count}")
    print(f"  Workouts: {workout_count}")
    print(f"  Exercises: {exercise_count}")
    print(f"  Exercise-Workout links: {link_count}")

    db.close()
