#!/usr/bin/env python3
"""Parser for prehab for performance-adv.xlsx (PFP Advanced: F2/R2/H2/Tx/Kinstretch).

Single column block (cols 2-6): time/section | label (A1/B1) | name | sets (xN) | reps.
Each workout starts with a title + W# header and runs WARM UP -> PRE-STRENGTH ->
STRENGTH -> STRETCH; a new "WARM UP" section after exercises = the next workout.

Letters REPEAT within a workout (two separate A-supersets), so groups are split
whenever the label number resets to 1, then relabelled sequentially A, B, C...

  --dump "<sheet>" | --import
"""

import os
import re
import sys
import sqlite3

import openpyxl

DB_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'ageless.db')
DATA_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'import-data')
FILE = 'prehab for performance-adv.xlsx'
PREFIX = 'PFP Advanced'

C_TIME, C_LABEL, C_NAME, C_SETS, C_REPS = 2, 3, 4, 5, 6
LABEL_RE = re.compile(r'^([A-Z])(\d+)$')
SECTIONS = ('WARM UP', 'PRE-STRENGTH', 'STRENGTH', 'STRETCH', 'MOBILITY', 'CONDITIONING', 'PRE STRENGTH', 'SKILL')
LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'


def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return '' if v is None else re.sub(r'\s+', ' ', str(v)).strip()


def parse_reps(raw):
    out = {'reps': None, 'duration_secs': None, 'time_based': 0, 'per_side': None, 'tracking_type': None}
    if not raw:
        return out
    val = raw.strip()
    if re.search(r'e\s*[./]\s*s|each side|per side', val, re.IGNORECASE):
        out['per_side'] = 'side'
    clean = re.sub(r'\s*(e\s*[./]\s*s\.?|each side|per side)', '', val, flags=re.IGNORECASE).strip()
    ms = re.match(r'^(\d+)\s*s$', clean, re.IGNORECASE)
    mt = re.match(r'^(\d+):(\d{2})$', clean)
    if mt:
        out.update(duration_secs=int(mt.group(1)) * 60 + int(mt.group(2)), time_based=1, tracking_type='Duration', reps=clean)
        return out
    if ms:
        out.update(duration_secs=int(ms.group(1)), time_based=1, tracking_type='Duration', reps=clean)
        return out
    mr = re.match(r'^(\d+(?:\s*-\s*\d+)?)\s*r', clean, re.IGNORECASE)
    if mr:
        out.update(reps=re.sub(r'\s*', '', mr.group(1)), tracking_type='Repetitions')
        return out
    out['reps'] = clean
    return out


def parse_sheet(ws):
    workouts = []
    cur = None
    section = ''
    cur_group = None      # list of exercise dicts in the active group
    seen_ex = False

    def new_workout():
        nonlocal cur, cur_group, seen_ex
        cur = {'groups': []}
        workouts.append(cur)
        cur_group = None
        seen_ex = False

    for r in range(1, ws.max_row + 1):
        timev = cell(ws, r, C_TIME)
        labelv = cell(ws, r, C_LABEL)
        namev = cell(ws, r, C_NAME)
        setsv = cell(ws, r, C_SETS)
        repsv = cell(ws, r, C_REPS)

        up = timev.upper()
        if up in SECTIONS or any(up == s for s in SECTIONS):
            if up == 'WARM UP' and seen_ex:
                new_workout()
            section = up
            continue

        lm = LABEL_RE.match(labelv)
        if not (lm and namev and namev != '-'):
            continue
        if cur is None:
            new_workout()
        letter, num = lm.group(1), int(lm.group(2))
        # new group when the item number resets to 1 (or no group yet)
        if num == 1 or cur_group is None:
            cur_group = {'items': [], 'section': section}
            cur['groups'].append(cur_group)
        sm = re.search(r'(\d+)', setsv)
        sets = int(sm.group(1)) if (sm and setsv.lower().startswith('x')) else None
        if sets:
            cur_group['sets'] = sets
        cur_group['items'].append({'name': namev, **parse_reps(repsv)})
        seen_ex = True

    # finalise: sequential group letters + type by size (warmup section override)
    result = []
    for wk in workouts:
        groups = [g for g in wk['groups'] if g['items']]
        if not groups:
            continue
        flat = []
        for gi, g in enumerate(groups):
            n = len(g['items'])
            warm = 'WARM' in g.get('section', '')
            gtype = 'warmup' if warm else ('superset' if n == 2 else 'triset' if n == 3 else 'circuit' if n >= 4 else 'regular')
            label = LETTERS[gi] if gi < len(LETTERS) else f'G{gi}'
            sets = g.get('sets', 1)
            for e in g['items']:
                e.update(group_label=label, group_type=gtype, sets=sets)
                flat.append(e)
        result.append(flat)
    return result


def _norm(s):
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9 ]', '', s.lower())).strip()


def dump(sheet):
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, FILE), data_only=True)
    workouts = parse_sheet(wb[sheet])
    print(f"{sheet}: {len(workouts)} workouts")
    for wi, wk in enumerate(workouts[:2], 1):
        print(f"\n  -- Workout {wi} ({len(wk)} ex) --")
        last = None
        for e in wk:
            if e['group_label'] != last:
                print(f"    [{e['group_label']}] {e['group_type'].upper()} x{e['sets']}")
                last = e['group_label']
            val = f"{e['duration_secs']}s" if e['time_based'] and e['duration_secs'] else (e['reps'] or '-')
            side = f" /{e['per_side']}" if e['per_side'] else ''
            print(f"        - {e['name'][:42]}  ({val}{side})")


def run_import():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    idx = {_norm(r['name']): r['id'] for r in db.execute('SELECT id, name FROM exercises').fetchall()}

    def find_or_create(name):
        k = _norm(name)
        if k in idx:
            return idx[k]
        c = db.execute('INSERT INTO exercises (name) VALUES (?)', (name,))
        idx[k] = c.lastrowid
        return c.lastrowid

    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, FILE), data_only=True)
    for sheet in wb.sheetnames:
        title = f"{PREFIX} | {sheet}"
        prog = db.execute('SELECT id FROM programs WHERE title = ?', (title,)).fetchone()
        if not prog:
            print(f"  SKIP (no program): {title}")
            continue
        pid = prog['id']
        workouts = parse_sheet(wb[sheet])
        if not workouts:
            print(f"  {title}: nothing parsed")
            continue
        old = [r['id'] for r in db.execute('SELECT id FROM workouts WHERE program_id=?', (pid,)).fetchall()]
        for wid in old:
            for x in db.execute('SELECT id FROM workout_exercises WHERE workout_id=?', (wid,)).fetchall():
                db.execute('DELETE FROM workout_exercise_meta WHERE workout_exercise_id=?', (x['id'],))
            db.execute('DELETE FROM workout_exercises WHERE workout_id=?', (wid,))
        db.execute('DELETE FROM workouts WHERE program_id=?', (pid,))

        wpw = 2
        db.execute('UPDATE programs SET duration_weeks=?, workouts_per_week=? WHERE id=?',
                   (max(1, (len(workouts) + wpw - 1) // wpw), wpw, pid))
        ex_total = 0
        for i, wk in enumerate(workouts):
            week, day = i // wpw + 1, i % wpw + 1
            c = db.execute(
                """INSERT INTO workouts (program_id, week_number, day_number, title, duration_mins, intensity, body_parts, workout_type, status)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (pid, week, day, f"{sheet} W{week} - Session {day}", 40, 'Medium', sheet, 'strength', 'draft'))
            wid = c.lastrowid
            for order, e in enumerate(wk):
                ex_id = find_or_create(e['name'])
                c2 = db.execute(
                    """INSERT INTO workout_exercises
                       (workout_id, exercise_id, order_index, sets, reps, duration_secs, rest_secs, group_type, group_label)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (wid, ex_id, order, e['sets'], e['reps'] if e['reps'] is not None else '',
                     e['duration_secs'], 30, e['group_type'], e['group_label']))
                db.execute(
                    """INSERT INTO workout_exercise_meta (workout_exercise_id, per_side, time_based, duration_secs, tracking_type)
                       VALUES (?,?,?,?,?)""",
                    (c2.lastrowid, e['per_side'], e['time_based'], e['duration_secs'], e['tracking_type']))
                ex_total += 1
        db.commit()
        print(f"  {title} (id={pid}): {len(workouts)} workouts, {ex_total} exercises")
    db.execute("DELETE FROM explore_section_items WHERE item_type='workout' AND NOT EXISTS (SELECT 1 FROM workouts w WHERE w.id=explore_section_items.item_id)")
    db.commit()
    db.close()


if __name__ == '__main__':
    if len(sys.argv) >= 3 and sys.argv[1] == '--dump':
        dump(sys.argv[2])
    elif len(sys.argv) >= 2 and sys.argv[1] == '--import':
        run_import()
    else:
        print('usage: parse_prehab_adv.py [--dump "<sheet>" | --import]')
